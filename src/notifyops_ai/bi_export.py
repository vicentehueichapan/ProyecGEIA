from __future__ import annotations

from pathlib import Path
from typing import Mapping

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


TITLE_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
SECTION_FILL = PatternFill("solid", fgColor="E0F2FE")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2EC"),
    right=Side(style="thin", color="D9E2EC"),
    top=Side(style="thin", color="D9E2EC"),
    bottom=Side(style="thin", color="D9E2EC"),
)


def build_security_audit_table() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "asset": "event_id",
                "sensitivity": "baja",
                "risk": "identificador tecnico reutilizable",
                "control": "mantenerlo como id tecnico sin datos personales directos",
                "roles": "DataOps, Analista BI, Docente evaluador",
                "law_alignment": "minimizacion y finalidad de tratamiento",
            },
            {
                "asset": "source_user_id",
                "sensitivity": "media",
                "risk": "puede asociarse a actividad de una persona usuaria",
                "control": "pseudonimizar, restringir acceso y evitar exponerlo en graficos publicos",
                "roles": "DataOps y auditor autorizado",
                "law_alignment": "proteccion de datos personales Ley 19.628",
            },
            {
                "asset": "target_user_id",
                "sensitivity": "media",
                "risk": "permite inferir relaciones e interacciones entre usuarios",
                "control": "pseudonimizar, mostrar agregados en BI y aplicar acceso por rol",
                "roles": "DataOps y auditor autorizado",
                "law_alignment": "proporcionalidad y acceso limitado",
            },
            {
                "asset": "content",
                "sensitivity": "alta",
                "risk": "puede contener texto personal o informacion sensible escrita por usuarios",
                "control": "limitar almacenamiento, no mostrar texto completo en BI y revisar retencion",
                "roles": "solo auditor autorizado si existe justificacion",
                "law_alignment": "minimizacion, seguridad y finalidad",
            },
            {
                "asset": "created_at",
                "sensitivity": "media",
                "risk": "metadato de comportamiento y trazabilidad de actividad",
                "control": "usar para metricas operativas; publicar preferentemente datos agregados",
                "roles": "DataOps, Analista BI",
                "law_alignment": "uso proporcional al objetivo del sistema",
            },
            {
                "asset": "model_metrics",
                "sensitivity": "baja",
                "risk": "no contiene datos personales, pero describe rendimiento operacional",
                "control": "compartir como evidencia tecnica y mantener versionamiento",
                "roles": "Equipo proyecto, profesor, analistas",
                "law_alignment": "sin dato personal directo",
            },
            {
                "asset": "logs",
                "sensitivity": "media",
                "risk": "pueden incluir rutas locales, tiempos de ejecucion o trazas tecnicas",
                "control": "rotar logs, evitar secretos y revisar antes de publicar",
                "roles": "DataOps y administrador",
                "law_alignment": "seguridad del tratamiento",
            },
        ]
    )


def build_access_roles_table() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "role": "Administrador DataOps",
                "allowed_access": "pipeline, logs, datos procesados, modelo y configuracion",
                "restriction": "no publicar datos identificables sin necesidad",
            },
            {
                "role": "Analista BI",
                "allowed_access": "metricas, agregados, dashboard y resultados del modelo",
                "restriction": "sin acceso a contenido textual sensible completo",
            },
            {
                "role": "Auditor/Profesor evaluador",
                "allowed_access": "evidencias, metricas, README, informe y dashboard",
                "restriction": "solo fines academicos y revision tecnica",
            },
            {
                "role": "Usuario negocio",
                "allowed_access": "indicadores agregados y alertas",
                "restriction": "sin acceso a ids de usuarios ni contenido individual",
            },
        ]
    )


def build_powerbi_blueprint_table() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "page": "Rendimiento IA",
                "visual": "Tarjetas KPI",
                "source_sheet": "metricas_modelo",
                "fields": "accuracy, precision, recall, f1_score, roc_auc, gini",
                "purpose": "demostrar rendimiento del modelo solicitado en la rubrica",
            },
            {
                "page": "Rendimiento IA",
                "visual": "Matriz de confusion",
                "source_sheet": "matriz_confusion",
                "fields": "real_class, pred_valido, pred_riesgoso",
                "purpose": "explicar aciertos, falsos positivos y falsos negativos",
            },
            {
                "page": "Decision operacional",
                "visual": "Grafico de barras",
                "source_sheet": "decisiones_finales",
                "fields": "final_decision",
                "purpose": "comparar rechazados por reglas, revision IA y aprobados",
            },
            {
                "page": "Variables y calidad",
                "visual": "Top variables",
                "source_sheet": "pesos_variables",
                "fields": "feature, abs_weight",
                "purpose": "justificar que variables influyen en el filtro inteligente",
            },
            {
                "page": "Seguridad",
                "visual": "Tabla de auditoria",
                "source_sheet": "auditoria_seguridad",
                "fields": "asset, sensitivity, risk, control, roles",
                "purpose": "defender proteccion de datos sensibles y roles",
            },
        ]
    )


def _python_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, (np.integer, np.floating)):
        return value.item()
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def _style_header(row) -> None:
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="111827")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 42)


def _write_dataframe(wb: Workbook, sheet_name: str, frame: pd.DataFrame, table_name: str) -> None:
    ws = wb.create_sheet(sheet_name)
    if frame.empty:
        ws.append(["sin_datos"])
        ws.append(["Ejecutar pipeline para regenerar esta hoja."])
        _autosize(ws)
        return

    ws.append([str(column) for column in frame.columns])
    for _, row in frame.iterrows():
        ws.append([_python_value(value) for value in row.tolist()])

    _style_header(ws[1])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    table_ref = ws.dimensions
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    _autosize(ws)


def _pipeline_kpis_as_table(pipeline_kpis: pd.DataFrame | None) -> pd.DataFrame:
    if pipeline_kpis is None or pipeline_kpis.empty:
        return pd.DataFrame(
            [
                {
                    "metric": "sin_datos",
                    "value": "Ejecutar python -m src.notifyops.pipeline antes de regenerar Excel BI.",
                }
            ]
        )
    if len(pipeline_kpis) == 1 and "metric" not in pipeline_kpis.columns:
        row = pipeline_kpis.iloc[0]
        return pd.DataFrame([{"metric": column, "value": row[column]} for column in pipeline_kpis.columns])
    return pipeline_kpis.copy()


def export_powerbi_workbook(
    output_path: Path,
    metrics: Mapping[str, float],
    confusion_matrix: np.ndarray,
    quality_summary: pd.DataFrame,
    feature_weights: pd.DataFrame,
    final_decisions: pd.DataFrame,
    new_event_predictions: pd.DataFrame,
    pipeline_kpis: pd.DataFrame | None = None,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "resumen_bi"

    ws.merge_cells("A1:H1")
    ws["A1"] = "NotifyOps - Fuente BI Parcial 3"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=15)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Metricas del modelo"
    ws["A3"].fill = SECTION_FILL
    ws["A3"].font = Font(bold=True)
    metric_rows = ["accuracy", "precision", "recall", "f1_score", "roc_auc", "gini"]
    ws.append(["metric", "value"])
    for metric in metric_rows:
        ws.append([metric, float(metrics.get(metric, 0))])
    _style_header(ws[4])

    decision_counts = final_decisions["final_decision"].value_counts().reset_index()
    decision_counts.columns = ["final_decision", "count"]
    ws["D3"] = "Decision final reglas + IA"
    ws["D3"].fill = SECTION_FILL
    ws["D3"].font = Font(bold=True)
    ws["D4"] = "final_decision"
    ws["E4"] = "count"
    _style_header(ws[4][3:5])
    for row_idx, row in enumerate(decision_counts.itertuples(index=False), start=5):
        ws.cell(row=row_idx, column=4, value=row.final_decision)
        ws.cell(row=row_idx, column=5, value=int(row.count))

    ws["A12"] = "Uso en Power BI"
    ws["A12"].fill = SECTION_FILL
    ws["A12"].font = Font(bold=True)
    ws["A13"] = "Importar este archivo Excel y crear paginas con KPI del modelo, matriz de confusion, decisiones finales, variables relevantes y auditoria de seguridad."
    ws.merge_cells("A13:H14")
    ws["A13"].alignment = Alignment(wrap_text=True, vertical="top")

    metrics_chart = BarChart()
    metrics_chart.title = "Metricas IA"
    metrics_chart.y_axis.title = "Valor"
    metrics_chart.x_axis.title = "Metrica"
    metrics_chart.add_data(Reference(ws, min_col=2, min_row=4, max_row=10), titles_from_data=True)
    metrics_chart.set_categories(Reference(ws, min_col=1, min_row=5, max_row=10))
    metrics_chart.height = 7
    metrics_chart.width = 14
    ws.add_chart(metrics_chart, "A16")

    decision_chart = BarChart()
    decision_chart.title = "Decision final"
    decision_chart.y_axis.title = "Eventos"
    decision_chart.add_data(Reference(ws, min_col=5, min_row=4, max_row=4 + len(decision_counts)), titles_from_data=True)
    decision_chart.set_categories(Reference(ws, min_col=4, min_row=5, max_row=4 + len(decision_counts)))
    decision_chart.height = 7
    decision_chart.width = 14
    ws.add_chart(decision_chart, "D16")

    _autosize(ws)

    metrics_df = pd.DataFrame([metrics])
    matrix_df = pd.DataFrame(
        confusion_matrix,
        index=["real_valido", "real_riesgoso"],
        columns=["pred_valido", "pred_riesgoso"],
    ).reset_index(names="real_class")
    decisions_summary = decision_counts

    _write_dataframe(wb, "metricas_modelo", metrics_df, "MetricasModelo")
    _write_dataframe(wb, "matriz_confusion", matrix_df, "MatrizConfusion")
    _write_dataframe(wb, "calidad_datos", quality_summary, "CalidadDatos")
    _write_dataframe(wb, "pesos_variables", feature_weights, "PesosVariables")
    _write_dataframe(wb, "decisiones_finales", final_decisions, "DecisionesFinales")
    _write_dataframe(wb, "resumen_decisiones", decisions_summary, "ResumenDecisiones")
    _write_dataframe(wb, "predicciones_nuevas", new_event_predictions, "PrediccionesNuevas")
    _write_dataframe(wb, "rendimiento_local", _pipeline_kpis_as_table(pipeline_kpis), "RendimientoLocal")
    _write_dataframe(wb, "auditoria_seguridad", build_security_audit_table(), "AuditoriaSeguridad")
    _write_dataframe(wb, "roles_acceso", build_access_roles_table(), "RolesAcceso")
    _write_dataframe(wb, "guia_powerbi", build_powerbi_blueprint_table(), "GuiaPowerBI")

    wb.save(output_path)
    return output_path
