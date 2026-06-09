from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data"
AI_REPORTS = DATA / "reports" / "ai"
BI_DATA = DATA / "bi"
PIPELINE_REPORTS = DATA / "reports"
DEFAULT_OUTPUT = BI_DATA / "notifyops_powerbi_dataset.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="17324D")
SECTION_FILL = PatternFill("solid", fgColor="DCEAF5")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=18, bold=True, color="17324D")


def _read_csv(path: Path, required: bool = True) -> pd.DataFrame:
    if path.exists():
        return pd.read_csv(path)
    if required:
        raise FileNotFoundError(f"Falta artefacto requerido para BI: {path}")
    return pd.DataFrame()


def pseudonymize_identifier(value: object) -> str:
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    digest = hashlib.sha256(f"notifyops-academic-salt:{text}".encode("utf-8")).hexdigest()
    return f"usr_{digest[:10]}"


def _security_audit() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "asset": "event_id",
                "sensitivity": "baja",
                "risk": "reutilizacion del identificador tecnico",
                "implemented_control": "identificador tecnico sin datos personales directos",
                "recommended_control": "rotacion o tokenizacion si se integra con sistemas externos",
                "authorized_roles": "DataOps, Analista BI, Auditor",
                "legal_alignment": "finalidad y minimizacion",
            },
            {
                "asset": "source_user_id",
                "sensitivity": "media",
                "risk": "asociacion de actividad a una persona usuaria",
                "implemented_control": "seudonimizacion SHA-256 antes de exportar a BI",
                "recommended_control": "separar tabla de correspondencias y limitar su acceso",
                "authorized_roles": "Administrador DataOps",
                "legal_alignment": "Ley 19.628: acceso limitado y uso proporcional",
            },
            {
                "asset": "target_user_id",
                "sensitivity": "media",
                "risk": "inferencia de relaciones entre usuarios",
                "implemented_control": "seudonimizacion y visualizacion agregada",
                "recommended_control": "aplicar retencion limitada y auditoria de acceso",
                "authorized_roles": "Administrador DataOps",
                "legal_alignment": "Ley 19.628: seguridad y finalidad",
            },
            {
                "asset": "content",
                "sensitivity": "alta",
                "risk": "texto libre con posibles datos personales o sensibles",
                "implemented_control": "contenido excluido del dataset BI",
                "recommended_control": "cifrado y acceso excepcional en una implementacion productiva",
                "authorized_roles": "Auditor autorizado con justificacion",
                "legal_alignment": "minimizacion y tratamiento limitado",
            },
            {
                "asset": "created_at",
                "sensitivity": "media",
                "risk": "perfilamiento temporal de actividad",
                "implemented_control": "uso operacional y filtros; no identifica por si solo",
                "recommended_control": "agregacion temporal para reportes publicos",
                "authorized_roles": "DataOps, Analista BI, Auditor",
                "legal_alignment": "proporcionalidad",
            },
            {
                "asset": "logs",
                "sensitivity": "media",
                "risk": "exposicion de rutas, errores o trazas tecnicas",
                "implemented_control": "logs sin contrasenas ni contenido de usuarios",
                "recommended_control": "rotacion, retencion y acceso por rol",
                "authorized_roles": "Administrador DataOps, Auditor",
                "legal_alignment": "seguridad del tratamiento",
            },
            {
                "asset": "model_metrics",
                "sensitivity": "baja",
                "risk": "interpretacion incorrecta del rendimiento",
                "implemented_control": "versionamiento junto con fecha, particion y modelo seleccionado",
                "recommended_control": "monitoreo de drift y recalibracion",
                "authorized_roles": "Equipo, Profesor, Analista BI",
                "legal_alignment": "sin dato personal directo",
            },
        ]
    )


def _roles() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "role": "Administrador DataOps",
                "allowed_access": "pipeline, logs, datos tecnicos, modelo y configuracion",
                "restriction": "no publicar datos identificables sin necesidad",
            },
            {
                "role": "Analista BI",
                "allowed_access": "metricas, agregados, claves seudonimizadas y dashboard",
                "restriction": "sin ids originales ni contenido textual",
            },
            {
                "role": "Auditor/Profesor",
                "allowed_access": "evidencias, metricas, codigo, informe y dashboard",
                "restriction": "uso academico y revision tecnica",
            },
            {
                "role": "Usuario de negocio",
                "allowed_access": "indicadores agregados y decisiones operacionales",
                "restriction": "sin datos individuales ni configuracion del modelo",
            },
        ]
    )


def _guide() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "page": "Resumen ejecutivo",
                "visual": "Tarjetas KPI",
                "source_sheet": "metricas_modelo",
                "fields": "accuracy, precision, recall, f1_score, roc_auc, gini",
                "interaction": "filtros por evento, decision y fecha",
                "purpose": "resumir rendimiento del modelo seleccionado",
            },
            {
                "page": "Resumen ejecutivo",
                "visual": "Barras de decisiones",
                "source_sheet": "resumen_decisiones",
                "fields": "final_decision, count",
                "interaction": "seleccion cruzada con tabla de eventos",
                "purpose": "mostrar rechazo, revision IA y aprobacion",
            },
            {
                "page": "Modelo y calidad",
                "visual": "Comparacion de modelos",
                "source_sheet": "comparacion_modelos",
                "fields": "model, f1_score, roc_auc, recall, selected",
                "interaction": "seleccion por modelo",
                "purpose": "justificar algoritmo final",
            },
            {
                "page": "Modelo y calidad",
                "visual": "Matriz de confusion",
                "source_sheet": "matriz_confusion",
                "fields": "real_class, pred_valido, pred_riesgoso",
                "interaction": "tooltip con interpretacion",
                "purpose": "explicar aciertos y errores",
            },
            {
                "page": "Modelo y calidad",
                "visual": "Curva ROC",
                "source_sheet": "curva_roc",
                "fields": "false_positive_rate, true_positive_rate",
                "interaction": "tooltip por punto",
                "purpose": "mostrar capacidad de separacion",
            },
            {
                "page": "Seguridad y operacion",
                "visual": "Matriz de auditoria",
                "source_sheet": "auditoria_seguridad",
                "fields": "asset, sensitivity, risk, implemented_control, authorized_roles",
                "interaction": "filtro por sensibilidad",
                "purpose": "defender proteccion y roles",
            },
        ]
    )


def _sanitize_decisions(decisions: pd.DataFrame) -> pd.DataFrame:
    sanitized = decisions.copy()
    sanitized["source_user_key"] = sanitized["source_user_id"].map(pseudonymize_identifier)
    sanitized["target_user_key"] = sanitized["target_user_id"].map(pseudonymize_identifier)
    return sanitized.drop(columns=["source_user_id", "target_user_id"], errors="ignore")


def _autosize(worksheet, max_width: int = 48) -> None:
    for column_cells in worksheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        width = min(max((len(value) for value in values), default=8) + 2, max_width)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(width, 11)


def _add_dataframe_sheet(workbook: Workbook, name: str, frame: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(name)
    headers = [str(column) for column in frame.columns]
    if not headers:
        worksheet.append(["sin_datos"])
        worksheet.append(["No existen datos para esta ejecucion."])
        _autosize(worksheet)
        return

    worksheet.append(headers)
    for row in frame.itertuples(index=False, name=None):
        worksheet.append([None if pd.isna(value) else value for value in row])

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    if len(frame) > 0:
        table_name = f"tbl_{name}"[:31]
        table = Table(displayName=table_name, ref=worksheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    _autosize(worksheet)


def _build_summary_sheet(
    workbook: Workbook,
    metrics: pd.DataFrame,
    decisions_summary: pd.DataFrame,
    metadata: dict[str, object],
) -> None:
    worksheet = workbook.create_sheet("resumen_bi")
    worksheet["A1"] = "NotifyOps - Fuente BI Parcial 3"
    worksheet["A1"].font = TITLE_FONT
    worksheet.merge_cells("A1:F1")
    worksheet["A3"] = "Modelo seleccionado"
    worksheet["B3"] = metadata.get("model", "")
    worksheet["A4"] = "Generado UTC"
    worksheet["B4"] = metadata.get("generated_at_utc", "")

    worksheet["A6"] = "Metrica"
    worksheet["B6"] = "Valor"
    for cell in worksheet[6]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
    metric_names = ["accuracy", "precision", "recall", "f1_score", "roc_auc", "gini"]
    for row_index, metric in enumerate(metric_names, start=7):
        worksheet.cell(row=row_index, column=1, value=metric)
        worksheet.cell(row=row_index, column=2, value=float(metrics.iloc[0][metric]))
        worksheet.cell(row=row_index, column=2).number_format = "0.00%"

    worksheet["D6"] = "Decision final"
    worksheet["E6"] = "Cantidad"
    worksheet["D6"].fill = HEADER_FILL
    worksheet["E6"].fill = HEADER_FILL
    worksheet["D6"].font = WHITE_FONT
    worksheet["E6"].font = WHITE_FONT
    for row_index, row in enumerate(decisions_summary.itertuples(index=False), start=7):
        worksheet.cell(row=row_index, column=4, value=str(row.final_decision).replace("_", " "))
        worksheet.cell(row=row_index, column=5, value=int(row.count))

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Decisiones finales"
    chart.x_axis.title = "Cantidad"
    chart.y_axis.title = "Decision"
    chart.legend = None
    chart.add_data(Reference(worksheet, min_col=5, min_row=6, max_row=6 + len(decisions_summary)), titles_from_data=True)
    chart.set_categories(Reference(worksheet, min_col=4, min_row=7, max_row=6 + len(decisions_summary)))
    chart.height = 7
    chart.width = 13
    worksheet.add_chart(chart, "G3")

    worksheet["A15"] = "Uso"
    worksheet["A15"].fill = SECTION_FILL
    worksheet["A15"].font = Font(bold=True, color="17324D")
    worksheet["A16"] = (
        "Importar este archivo en Power BI para construir las paginas Resumen ejecutivo, "
        "Modelo y calidad, y Seguridad y operacion. Las hojas contienen datos tabulares "
        "generados desde la misma ejecucion del pipeline."
    )
    worksheet.merge_cells("A16:F18")
    worksheet["A16"].alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(worksheet)


def build_powerbi_workbook(output_path: Path = DEFAULT_OUTPUT) -> Path:
    metrics = _read_csv(AI_REPORTS / "model_metrics.csv")
    comparison = _read_csv(AI_REPORTS / "model_comparison.csv")
    confusion = _read_csv(AI_REPORTS / "confusion_matrix.csv")
    roc_points = _read_csv(AI_REPORTS / "roc_curve_points.csv")
    quality = _read_csv(AI_REPORTS / "quality_summary.csv")
    weights = _read_csv(AI_REPORTS / "feature_weights.csv")
    decisions = _read_csv(AI_REPORTS / "final_event_decisions.csv")
    predictions = _read_csv(AI_REPORTS / "new_event_predictions.csv")
    performance_ai = _read_csv(AI_REPORTS / "performance_summary.csv")
    performance_local = _read_csv(PIPELINE_REPORTS / "kpi_report.csv", required=False)

    statistics_mask = quality["indicator"].astype(str).str.contains(
        r"_(?:mean|median|mode|p25|p75)$",
        regex=True,
    )
    quality_statistics = quality[statistics_mask].reset_index(drop=True)
    quality_operational = quality[~statistics_mask].reset_index(drop=True)
    decisions_summary = (
        decisions["final_decision"].value_counts().rename_axis("final_decision").reset_index(name="count")
    )
    sanitized_decisions = _sanitize_decisions(decisions)

    model_metadata_path = ROOT / "models" / "notifyops_ai_model.json"
    metadata = json.loads(model_metadata_path.read_text(encoding="utf-8")) if model_metadata_path.exists() else {}

    workbook = Workbook()
    workbook.remove(workbook.active)
    _build_summary_sheet(workbook, metrics, decisions_summary, metadata)
    sheets: Iterable[tuple[str, pd.DataFrame]] = [
        ("metricas_modelo", metrics),
        ("comparacion_modelos", comparison),
        ("matriz_confusion", confusion),
        ("curva_roc", roc_points),
        ("calidad_datos", quality_operational),
        ("estadisticas_calidad", quality_statistics),
        ("pesos_variables", weights),
        ("decisiones_finales", sanitized_decisions),
        ("resumen_decisiones", decisions_summary),
        ("predicciones_nuevas", predictions),
        ("rendimiento_local", performance_local),
        ("rendimiento_ia", performance_ai),
        ("auditoria_seguridad", _security_audit()),
        ("roles_acceso", _roles()),
        ("guia_powerbi", _guide()),
    ]
    for name, frame in sheets:
        _add_dataframe_sheet(workbook, name, frame)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


if __name__ == "__main__":
    print(build_powerbi_workbook())
