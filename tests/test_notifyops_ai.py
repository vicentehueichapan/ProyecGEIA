import unittest
import zipfile
from pathlib import Path
from xml.etree import ElementTree

import pandas as pd
from openpyxl import load_workbook

from src.notifyops_ai import modeling
from src.notifyops_ai import bi_dataset


class NotifyOpsAITests(unittest.TestCase):
    def test_synthetic_dataset_contains_valid_and_risky_events(self):
        events = modeling.generate_synthetic_events(rows=120, seed=7)

        self.assertEqual(len(events), 120)
        self.assertIn("label_risky_event", events.columns)
        self.assertGreater(events["label_risky_event"].sum(), 0)
        self.assertGreater((events["label_risky_event"] == 0).sum(), 0)

    def test_dataset_contains_predictive_feedback_risk_independent_of_hard_rules(self):
        events = modeling.generate_synthetic_events(rows=500, seed=42)
        clean_events = events[events.apply(modeling.rule_error_reason, axis=1).eq("")]

        self.assertGreater(clean_events["label_risky_event"].sum(), 0)
        self.assertTrue(
            {
                "interaction_velocity_5m",
                "account_age_days",
                "historical_report_rate",
            }.issubset(events.columns)
        )

    def test_feature_engineering_creates_model_columns(self):
        events = pd.DataFrame(
            [
                {
                    "event_id": "evt-001",
                    "event_type": "LIKE",
                    "source_user_id": "u01",
                    "target_user_id": "u02",
                    "created_at": "2026-05-14 09:00:00",
                    "content": "",
                    "is_duplicate": 0,
                    "label_risky_event": 0,
                },
                {
                    "event_id": "evt-002",
                    "event_type": "share",
                    "source_user_id": "u03",
                    "target_user_id": "",
                    "created_at": "fecha-invalida",
                    "content": "externo",
                    "is_duplicate": 0,
                    "label_risky_event": 1,
                },
            ]
        )

        engineered = modeling.engineer_features(events)

        for column in modeling.FEATURE_COLUMNS:
            self.assertIn(column, engineered.columns)
        self.assertEqual(engineered.loc[0, "is_allowed_event_type"], 1)
        self.assertEqual(engineered.loc[1, "event_type_invalid"], 1)
        self.assertEqual(engineered.loc[1, "has_target_user"], 0)
        self.assertEqual(engineered.loc[1, "has_valid_date"], 0)

    def test_ai_pipeline_produces_required_metrics(self):
        result = modeling.run_ai_pipeline(rows=160, seed=11, save_plots=False, write_outputs=False)

        for metric in ["accuracy", "precision", "recall", "f1_score", "roc_auc", "gini"]:
            self.assertIn(metric, result.metrics)
            self.assertGreaterEqual(result.metrics[metric], 0)
        self.assertEqual(result.confusion_matrix.shape, (2, 2))
        self.assertIn("feature", result.feature_weights.columns)

    def test_roc_calculation_supports_numpy_1_26_without_trapezoid(self):
        original_trapezoid = modeling.np.trapezoid
        had_trapz = hasattr(modeling.np, "trapz")
        original_trapz = getattr(modeling.np, "trapz", None)
        modeling.np.trapz = original_trapezoid
        delattr(modeling.np, "trapezoid")
        try:
            _, _, auc = modeling.roc_curve_values(
                modeling.np.array([0, 0, 1, 1]),
                modeling.np.array([0.1, 0.4, 0.35, 0.8]),
            )
        finally:
            modeling.np.trapezoid = original_trapezoid
            if had_trapz:
                modeling.np.trapz = original_trapz
            else:
                delattr(modeling.np, "trapz")

        self.assertAlmostEqual(auc, 0.75)

    def test_ai_pipeline_compares_models_and_records_runtime(self):
        result = modeling.run_ai_pipeline(rows=240, seed=42, save_plots=False, write_outputs=False)

        self.assertGreaterEqual(len(result.model_comparison), 3)
        self.assertIn("selected", result.model_comparison.columns)
        self.assertEqual(int(result.model_comparison["selected"].sum()), 1)
        self.assertGreater(result.performance["training_seconds"], 0)
        self.assertGreater(result.performance["inference_rows_per_second"], 0)

    def test_quality_analysis_contains_statistics_and_imputation_strategy(self):
        events = modeling.generate_synthetic_events(rows=200, seed=42)

        quality = modeling.quality_summary(events)
        indicators = set(quality["indicator"])

        for indicator in [
            "content_length_mean",
            "content_length_median",
            "content_length_mode",
            "content_length_p25",
            "content_length_p75",
            "interaction_velocity_5m_mean",
            "historical_report_rate_p75",
            "imputation_strategy",
        ]:
            self.assertIn(indicator, indicators)

    def test_final_decision_keeps_hard_rules_before_ai(self):
        events = modeling.generate_synthetic_events(rows=120, seed=21)
        engineered = modeling.engineer_features(events)
        x_train, x_test, y_train, y_test = modeling.stratified_train_test_split(
            engineered[modeling.FEATURE_COLUMNS],
            engineered["label_risky_event"].astype(int),
            seed=21,
        )
        x_train_scaled, _, mean, std = modeling.standardize_train_test(x_train, x_test)
        weights = modeling.train_logistic_regression(x_train_scaled, y_train.to_numpy(dtype=int))

        sample = pd.DataFrame(
            [
                {
                    "event_id": "evt-bad",
                    "event_type": "share",
                    "source_user_id": "u01",
                    "target_user_id": "",
                    "created_at": "fecha-invalida",
                    "content": "externo",
                    "is_duplicate": 0,
                    "label_risky_event": 1,
                }
            ]
        )

        decisions = modeling.build_final_decisions(sample, weights, mean, std)

        self.assertEqual(decisions.iloc[0]["final_decision"], "rechazado_por_reglas")
        self.assertIn("target_user_id vacio", decisions.iloc[0]["rule_error_reason"])
        self.assertIn("fecha invalida", decisions.iloc[0]["rule_error_reason"])

    def test_final_decisions_include_all_three_operational_states(self):
        modeling.run_ai_pipeline(rows=500, seed=42, save_plots=False, write_outputs=True)
        decisions = pd.read_csv("data/reports/ai/final_event_decisions.csv")

        self.assertEqual(
            set(decisions["final_decision"]),
            {"rechazado_por_reglas", "revision_por_ia", "aprobado_para_notificar"},
        )

    def test_powerbi_fixed_workbook_contains_required_sheets(self):
        workbook_path = Path("data/bi/notifyops_powerbi_dataset.xlsx")
        if workbook_path.exists():
            workbook_path.unlink()
        modeling.run_ai_pipeline(rows=240, seed=42, save_plots=False, write_outputs=True)

        self.assertTrue(workbook_path.exists())
        with zipfile.ZipFile(workbook_path) as workbook_zip:
            workbook_xml = workbook_zip.read("xl/workbook.xml")

        namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        root = ElementTree.fromstring(workbook_xml)
        sheet_names = {sheet.attrib["name"] for sheet in root.findall(".//main:sheet", namespace)}

        for sheet_name in [
            "resumen_bi",
            "metricas_modelo",
            "comparacion_modelos",
            "matriz_confusion",
            "curva_roc",
            "estadisticas_calidad",
            "decisiones_finales",
            "rendimiento_ia",
            "auditoria_seguridad",
            "guia_powerbi",
        ]:
            self.assertIn(sheet_name, sheet_names)

    def test_powerbi_workbook_matches_current_metrics_and_pseudonymizes_users(self):
        modeling.run_ai_pipeline(rows=240, seed=42, save_plots=False, write_outputs=True)
        workbook_path = bi_dataset.build_powerbi_workbook()
        csv_metrics = pd.read_csv("data/reports/ai/model_metrics.csv").iloc[0].to_dict()
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)

        metric_sheet = workbook["metricas_modelo"]
        headers = [cell.value for cell in next(metric_sheet.iter_rows(min_row=1, max_row=1))]
        values = [cell.value for cell in next(metric_sheet.iter_rows(min_row=2, max_row=2))]
        workbook_metrics = dict(zip(headers, values))
        self.assertAlmostEqual(workbook_metrics["accuracy"], csv_metrics["accuracy"], places=4)
        self.assertAlmostEqual(workbook_metrics["gini"], csv_metrics["gini"], places=4)

        decisions_sheet = workbook["decisiones_finales"]
        decision_headers = [cell.value for cell in next(decisions_sheet.iter_rows(min_row=1, max_row=1))]
        self.assertIn("source_user_key", decision_headers)
        self.assertIn("target_user_key", decision_headers)
        self.assertNotIn("source_user_id", decision_headers)
        self.assertNotIn("target_user_id", decision_headers)
        workbook.close()

    def test_dashboard_is_interactive_and_uses_generated_data(self):
        modeling.run_ai_pipeline(rows=240, seed=42, save_plots=True, write_outputs=True)
        html_path = Path("dashboard/notifyops_ai_dashboard.html")
        data_path = Path("dashboard/data/dashboard_data.json")

        self.assertTrue(data_path.exists())
        html = html_path.read_text(encoding="utf-8")
        self.assertIn('id="eventTypeFilter"', html)
        self.assertIn('id="decisionFilter"', html)
        self.assertIn('id="resetFilters"', html)
        self.assertIn("dashboard_data.json", html)
        self.assertIn("<canvas", html)
        self.assertIn("<script", html)
        self.assertTrue(Path("data/reports/ai/charts/model_comparison.png").exists())
        self.assertTrue(Path("data/reports/ai/charts/runtime_comparison.png").exists())


if __name__ == "__main__":
    unittest.main()
